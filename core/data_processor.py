from typing import Tuple, List
from utils.logger import log_error
from utils.validators import is_valid_range_format
import openpyxl
import re
from datetime import datetime

def column_to_index(column_str: str) -> int:
    result = 0
    for char in column_str:
        if char.isalpha():
            result = result * 26 + (ord(char.upper()) - ord('A') + 1)
    return result

def parse_range(range_str: str) -> Tuple[int, int, int, int]:
    # 단일 셀 자동 변환
    is_valid, converted_range = is_valid_range_format(range_str)
    if not is_valid:
        raise ValueError(f"Invalid range format: {range_str}")
    
    range_str = converted_range
    
    start_cell, end_cell = range_str.split(':')
    
    match_start = re.match(r'([A-Z]+)(\d+)', start_cell)
    if not match_start:
        raise ValueError(f"Invalid start cell format: {start_cell}")
    start_col_str = match_start.group(1)
    start_row = int(match_start.group(2)) - 1
    
    match_end = re.match(r'([A-Z]+)(\d+)', end_cell)
    if not match_end:
        raise ValueError(f"Invalid end cell format: {end_cell}")
    end_col_str = match_end.group(1)
    end_row = int(match_end.group(2)) - 1
    
    start_col = column_to_index(start_col_str) - 1
    end_col = column_to_index(end_col_str) - 1
    
    return (start_row, end_row, start_col, end_col)

def merge_cell_sum(wb1: openpyxl.Workbook, wb2: openpyxl.Workbook, 
                   sheet_name: str, data_range: str) -> None:
    try:
        start_row, end_row, start_col, end_col = parse_range(data_range)
        
        sheet1 = wb1[sheet_name]
        sheet2 = wb2[sheet_name]

        for row in range(start_row + 1, end_row + 2):
            for col in range(start_col + 1, end_col + 2):
                cell1_value = sheet1.cell(row=row, column=col).value
                cell2_value = sheet2.cell(row=row, column=col).value

                if isinstance(cell1_value, (int, float)) and isinstance(cell2_value, (int, float)):
                    sheet1.cell(row=row, column=col, value=cell1_value + cell2_value)
    except Exception as e:
        log_error(f"셀 합산 병합 실패: {str(e)}")
        raise

def merge_row_append(wb1: openpyxl.Workbook, wb2: openpyxl.Workbook,
                     sheet_name: str, data_range: str) -> None:
    try:
        start_row, end_row, start_col, end_col = parse_range(data_range)
        
        sheet1 = wb1[sheet_name]
        sheet2 = wb2[sheet_name]

        last_row_wb1 = sheet1.max_row
        
        row_offset = 0
        for row in range(start_row + 1, end_row + 2):
            for col in range(start_col + 1, end_col + 2):
                value = sheet2.cell(row=row, column=col).value
                if value is not None:
                    sheet1.cell(row=last_row_wb1 + 1 + row_offset, column=col, value=value)
            row_offset += 1

    except Exception as e:
        log_error(f"행 누적 병합 실패: {str(e)}")
        raise

def merge_cell_min(wb1: openpyxl.Workbook, wb2: openpyxl.Workbook, 
                   sheet_name: str, data_range: str) -> None:
    """두 파일의 셀 값을 비교하여 최소값을 선택"""
    try:
        from datetime import datetime
        
        start_row, end_row, start_col, end_col = parse_range(data_range)
        
        sheet1 = wb1[sheet_name]
        sheet2 = wb2[sheet_name]

        for row in range(start_row + 1, end_row + 2):
            for col in range(start_col + 1, end_col + 2):
                cell1_value = sheet1.cell(row=row, column=col).value
                cell2_value = sheet2.cell(row=row, column=col).value

                # 숫자 또는 datetime 비교
                if isinstance(cell1_value, (int, float, datetime)) and isinstance(cell2_value, (int, float, datetime)):
                    min_value = min(cell1_value, cell2_value)
                    sheet1.cell(row=row, column=col, value=min_value)
                # 하나만 값이 있는 경우
                elif cell2_value is not None:
                    sheet1.cell(row=row, column=col, value=cell2_value)
                    
    except Exception as e:
        log_error(f"최소값 선택 병합 실패: {str(e)}")
        raise

def merge_cell_max(wb1: openpyxl.Workbook, wb2: openpyxl.Workbook, 
                   sheet_name: str, data_range: str) -> None:
    """두 파일의 셀 값을 비교하여 최대값을 선택"""
    try:
        from datetime import datetime
        
        start_row, end_row, start_col, end_col = parse_range(data_range)
        
        sheet1 = wb1[sheet_name]
        sheet2 = wb2[sheet_name]

        for row in range(start_row + 1, end_row + 2):
            for col in range(start_col + 1, end_col + 2):
                cell1_value = sheet1.cell(row=row, column=col).value
                cell2_value = sheet2.cell(row=row, column=col).value

                # 숫자 또는 datetime 비교
                if isinstance(cell1_value, (int, float, datetime)) and isinstance(cell2_value, (int, float, datetime)):
                    max_value = max(cell1_value, cell2_value)
                    sheet1.cell(row=row, column=col, value=max_value)
                # 하나만 값이 있는 경우
                elif cell2_value is not None:
                    sheet1.cell(row=row, column=col, value=cell2_value)
                    
    except Exception as e:
        log_error(f"최대값 선택 병합 실패: {str(e)}")
        raise

def extract_range_data(sheet: openpyxl.worksheet.worksheet.Worksheet, 
                       start_row: int, end_row: int, 
                       start_col: int, end_col: int) -> List[List]:
    data = []
    
    for row in range(start_row + 1, end_row + 2):
        row_data = []
        for col in range(start_col + 1, end_col + 2):
            cell_value = sheet.cell(row=row, column=col).value
            row_data.append(cell_value)
        data.append(row_data)
    
    return data

def group_and_sum_by_key(data1: List[List], data2: List[List], 
                          key_col_index: int) -> List[List]:
    result_dict = {}

    # data1을 딕셔너리에 저장
    for row in data1:
        key = row[key_col_index]
        if key not in result_dict:
            result_dict[key] = list(row)
    
    # data2를 순회하며 키 체크
    for row in data2:
        key = row[key_col_index]
        if key in result_dict:
            for i in range(len(row)):
                if i == key_col_index:
                    continue
                if isinstance(result_dict[key][i], (int, float)) and isinstance(row[i], (int, float)):
                    result_dict[key][i] += row[i]
        else:
            result_dict[key] = list(row)
    
    return list(result_dict.values())

def write_data_to_range(sheet: openpyxl.worksheet.worksheet.Worksheet, 
                        data: List[List], 
                        start_row: int, start_col: int) -> None:
    for row_idx, row_data in enumerate(data):
        for col_idx, value in enumerate(row_data):
            sheet.cell(row=start_row + 1 + row_idx, column=start_col + 1 + col_idx).value = value

def write_index_column(sheet: openpyxl.worksheet.worksheet.Worksheet, 
                      index_col: int, 
                      start_row: int, 
                      row_count: int) -> None:
    for i in range(row_count):
        sheet.cell(row=start_row + 1 + i, column=index_col + 1).value = i + 1

def merge_key_based(wb1: openpyxl.Workbook, wb2: openpyxl.Workbook,
                    sheet_name: str, data_range: str, 
                    key_column: str, index_column: str) -> None:
    try:
        start_row, end_row, start_col, end_col = parse_range(data_range)
        
        key_col_relative = (column_to_index(key_column) - 1) - start_col
        index_col_index = column_to_index(index_column) - 1
        
        sheet1 = wb1[sheet_name]
        sheet2 = wb2[sheet_name]

        data_wb1 = extract_range_data(sheet1, start_row, end_row, start_col, end_col)
        data_wb2 = extract_range_data(sheet2, start_row, end_row, start_col, end_col)

        merged_data = group_and_sum_by_key(data_wb1, data_wb2, key_col_relative)

        for row in range(start_row + 1, end_row + 2):
            for col in range(start_col + 1, end_col + 2):
                sheet1.cell(row=row, column=col).value = None

        write_data_to_range(sheet1, merged_data, start_row, start_col)
        
        row_count = len(merged_data)
        write_index_column(sheet1, index_col_index, start_row, row_count)

    except Exception as e:
        log_error(f"키 기반 병합 실패: {str(e)}")
        raise